"""
Report Generator for Risk Adjustment Analysis
Generates PDF and Excel reports for stakeholders
"""

import json
from datetime import datetime
from typing import Dict, List
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportGenerator:
    """Generate risk adjustment reports in multiple formats"""
    
    def __init__(self):
        self.timestamp = datetime.now()
    
    def generate_text_report(self, report_data: Dict) -> str:
        """Generate text format report"""
        lines = []
        lines.append("=" * 80)
        lines.append("RISK ADJUSTMENT ANALYSIS REPORT")
        lines.append(f"Generated: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 80)
        
        # Patient Information
        lines.append("\n[PATIENT INFORMATION]")
        demographics = report_data.get("demographics", {})
        lines.append(f"Patient ID: {report_data.get('patient_id', 'N/A')}")
        lines.append(f"Age: {demographics.get('age', 'N/A')}")
        lines.append(f"Gender: {demographics.get('gender', 'N/A')}")
        lines.append(f"Insurance Model: {report_data.get('insurance_model', 'N/A')}")
        
        # Diagnoses
        raf_calc = report_data.get("raf_calculation", {})
        lines.append("\n[DIAGNOSES & HCC MAPPINGS]")
        for mapping in raf_calc.get("hcc_mappings", []):
            lines.append(f"  • {mapping['icd10']}: {mapping['hcc_name']} (HCC: {mapping['hcc_id']}, RAF: {mapping['raf_value']})")
        
        # Risk Assessment
        lines.append("\n[RISK ASSESSMENT]")
        lines.append(f"RAF Score: {raf_calc.get('raf_score', 0.0):.3f}")
        lines.append(f"Risk Level: {report_data.get('risk_level', 'Unknown')}")
        
        # Premium Calculation
        lines.append("\n[PREMIUM CALCULATION]")
        premium = report_data.get("premium_calculation", {})
        lines.append(f"Base Premium: ${premium.get('base_premium', 0):,.2f}")
        lines.append(f"Age Factor: {premium.get('age_factor', 1.0):.2f}x")
        lines.append(f"Gender Factor: {premium.get('gender_factor', 1.0):.2f}x")
        lines.append(f"RAF Multiplier: {premium.get('raf_multiplier', 1.0):.3f}x")
        lines.append(f"Adjusted Annual Premium: ${premium.get('adjusted_premium', 0):,.2f}")
        lines.append(f"Adjusted Monthly Premium: ${premium.get('monthly_premium', 0):,.2f}")
        
        # Hierarchy Issues
        if raf_calc.get("hierarchy_issues"):
            lines.append("\n[HIERARCHY ISSUES & ALERTS]")
            for issue in raf_calc["hierarchy_issues"]:
                lines.append(f"  ⚠ {issue['message']}")
        
        # Recommendations
        if report_data.get("recommendations"):
            lines.append("\n[CODER RECOMMENDATIONS]")
            for rec in report_data["recommendations"]:
                lines.append(f"  • {rec}")
        
        lines.append("\n" + "=" * 80)
        
        return "\n".join(lines)
    
    def generate_json_report(self, report_data: Dict, output_file: str = None) -> str:
        """Generate JSON format report"""
        if output_file is None:
            output_file = f"risk_report_{self.timestamp.strftime('%Y%m%d_%H%M%S')}.json"
        
        report_data["report_generated"] = self.timestamp.isoformat()
        
        with open(output_file, 'w') as f:
            json.dump(report_data, f, indent=2)
        
        return output_file
    
    def generate_csv_report(self, patient_list: List[Dict], output_file: str = None) -> str:
        """Generate CSV report for bulk patient analysis"""
        if output_file is None:
            output_file = f"bulk_risk_report_{self.timestamp.strftime('%Y%m%d_%H%M%S')}.csv"
        
        import csv
        
        fieldnames = [
            'patient_id', 'age', 'gender', 'insurance_model', 
            'raf_score', 'risk_level', 'adjusted_premium', 'monthly_premium',
            'hcc_count', 'hierarchy_issues', 'report_date'
        ]
        
        with open(output_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for patient in patient_list:
                raf_calc = patient.get("raf_calculation", {})
                premium = patient.get("premium_calculation", {})
                
                row = {
                    'patient_id': patient.get('patient_id'),
                    'age': patient.get('demographics', {}).get('age'),
                    'gender': patient.get('demographics', {}).get('gender'),
                    'insurance_model': patient.get('insurance_model'),
                    'raf_score': raf_calc.get('raf_score', 0),
                    'risk_level': patient.get('risk_level'),
                    'adjusted_premium': premium.get('adjusted_premium', 0),
                    'monthly_premium': premium.get('monthly_premium', 0),
                    'hcc_count': len(raf_calc.get('hcc_mappings', [])),
                    'hierarchy_issues': len(raf_calc.get('hierarchy_issues', [])),
                    'report_date': self.timestamp.isoformat()
                }
                writer.writerow(row)
        
        return output_file
    
    def generate_pdf_report(self, report_data: Dict, output_file: str = None) -> str:
        """Generate PDF report (requires reportlab)"""
        if not REPORTLAB_AVAILABLE:
            return self.generate_text_report(report_data)  # Fallback to text
        
        if output_file is None:
            output_file = f"risk_report_{self.timestamp.strftime('%Y%m%d_%H%M%S')}.pdf"
        
        doc = SimpleDocTemplate(output_file, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=6
        )
        
        # Title
        elements.append(Paragraph("Risk Adjustment Analysis Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Patient Info
        demographics = report_data.get("demographics", {})
        patient_info = [
            ["Patient ID:", report_data.get('patient_id', 'N/A')],
            ["Age:", str(demographics.get('age', 'N/A'))],
            ["Gender:", demographics.get('gender', 'N/A')],
            ["Insurance Model:", report_data.get('insurance_model', 'N/A')]
        ]
        
        t = Table(patient_info, colWidths=[2*inch, 2*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f0f8')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.3*inch))
        
        # HCC Mappings
        elements.append(Paragraph("HCC Mappings & RAF Scores", title_style))
        raf_calc = report_data.get("raf_calculation", {})
        hcc_data = [["ICD-10", "HCC ID", "Description", "RAF Score"]]
        for mapping in raf_calc.get("hcc_mappings", []):
            hcc_data.append([
                mapping['icd10'],
                mapping['hcc_id'],
                mapping['hcc_name'],
                f"{mapping['raf_value']:.3f}"
            ])
        
        t = Table(hcc_data, colWidths=[1.2*inch, 1*inch, 2.2*inch, 0.8*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(t)
        
        doc.build(elements)
        return output_file
    
    def generate_excel_report(self, patient_list: List[Dict], output_file: str = None) -> str:
        """Generate Excel report with multiple sheets"""
        if not OPENPYXL_AVAILABLE:
            return self.generate_csv_report(patient_list, output_file)
        
        if output_file is None:
            output_file = f"bulk_risk_report_{self.timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "Risk Adjustment Analysis Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        summary_data = [
            ["Total Patients", len(patient_list)],
            ["Report Generated", self.timestamp.isoformat()],
            ["Average RAF Score", sum(p.get('raf_calculation', {}).get('raf_score', 0) for p in patient_list) / len(patient_list) if patient_list else 0]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 2):
            ws_summary[f'A{row_idx}'] = label
            ws_summary[f'B{row_idx}'] = value
        
        # Detailed sheet
        ws_detail = wb.create_sheet("Detailed Analysis")
        headers = ['Patient ID', 'Age', 'Gender', 'Insurance', 'RAF Score', 'Risk Level', 'Annual Premium', 'Monthly Premium']
        ws_detail.append(headers)
        
        # Format header
        for cell in ws_detail[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
        
        # Add patient data
        for patient in patient_list:
            raf_calc = patient.get("raf_calculation", {})
            premium = patient.get("premium_calculation", {})
            
            ws_detail.append([
                patient.get('patient_id'),
                patient.get('demographics', {}).get('age'),
                patient.get('demographics', {}).get('gender'),
                patient.get('insurance_model'),
                raf_calc.get('raf_score', 0),
                patient.get('risk_level'),
                premium.get('adjusted_premium', 0),
                premium.get('monthly_premium', 0)
            ])
        
        # Adjust column widths
        for column in ws_detail.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                max_length = max(max_length, len(str(cell.value)))
            ws_detail.column_dimensions[column_letter].width = max_length + 2
        
        wb.save(output_file)
        return output_file
