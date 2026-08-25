"""Workbook-backed ICD-10 diagnosis matching for chart NLP results."""

from __future__ import annotations

import os
import re
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path
from typing import Dict, List


_WORD_PATTERN = re.compile(r"[a-z0-9]+")
_DEFAULT_WORKBOOK = Path(__file__).resolve().parents[1] / "CMS Models" / "section111_valid_icd10_october2025_0.xlsx"


def _normalize(value: str) -> str:
    return " ".join(_WORD_PATTERN.findall(str(value).lower()))


def _tokens(value: str) -> set[str]:
    return set(_normalize(value).split())


def _format_code(code: str) -> str:
    compact = str(code).strip().upper().replace(".", "")
    return f"{compact[:3]}.{compact[3:]}" if len(compact) > 3 else compact


class ICDDiagnosisMatcher:
    """Search the official ICD workbook for diagnosis phrases and codes."""

    def __init__(self, workbook_path: str | None = None) -> None:
        self.workbook_path = Path(workbook_path or os.getenv("ICD_DIAGNOSIS_FILE", _DEFAULT_WORKBOOK))
        self.rows: List[dict] = []
        self.token_index: Dict[str, set[int]] = {}
        self.available = False
        self._load()

    def _load(self) -> None:
        if not self.workbook_path.is_file():
            return
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
            sheet = workbook["Valid ICD10 FY2026 & NF Exclude"] if "Valid ICD10 FY2026 & NF Exclude" in workbook.sheetnames else workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip().upper() for value in next(rows)]
            header_index = {header: index for index, header in enumerate(headers)}
            code_index = header_index.get("CODE")
            short_index = header_index.get("SHORT DESCRIPTION (VALID ICD-10 FY2026)")
            long_index = header_index.get("LONG DESCRIPTION (VALID ICD-10 FY2026)")
            if code_index is None or short_index is None or long_index is None:
                return

            for row in rows:
                code = _format_code(row[code_index] or "")
                short = str(row[short_index] or "").strip()
                long = str(row[long_index] or "").strip()
                if not code or not short:
                    continue
                index = len(self.rows)
                self.rows.append({"code": code, "short_description": short, "description": long or short})
                for token in _tokens(f"{short} {long}"):
                    self.token_index.setdefault(token, set()).add(index)
            workbook.close()
            self.available = bool(self.rows)
        except Exception:
            self.rows = []
            self.token_index = {}

    def lookup_code(self, code: str) -> dict | None:
        normalized_code = _format_code(code)
        for row in self.rows:
            if row["code"] == normalized_code:
                return row.copy()
        return None

    def find_similar(self, diagnosis: str, limit: int = 5) -> List[dict]:
        if not self.available or not diagnosis.strip():
            return []
        query = _normalize(diagnosis)
        query_tokens = set(query.split())
        candidate_ids: set[int] = set()
        for token in query_tokens:
            candidate_ids.update(self.token_index.get(token, set()))
        if not candidate_ids:
            candidate_ids = set(range(len(self.rows)))

        matches = []
        for index in candidate_ids:
            row = self.rows[index]
            text = _normalize(f"{row['short_description']} {row['description']}")
            row_tokens = set(text.split())
            overlap = len(query_tokens & row_tokens) / max(len(query_tokens), 1)
            sequence = SequenceMatcher(None, query, text).ratio()
            score = min(1.0, (overlap * 0.7) + (sequence * 0.3))
            if score >= 0.35:
                matches.append({**row, "similarity": round(score, 3)})
        matches.sort(key=lambda item: (-item["similarity"], item["code"]))
        return matches[:limit]

    def search(self, query: str, limit: int = 100) -> List[dict]:
        query = query.strip().lower()
        if not query:
            return []
        matches = []
        for row in self.rows:
            searchable = f"{row['code']} {row['short_description']} {row['description']}".lower()
            if query in searchable:
                matches.append(row.copy())
                if len(matches) >= limit:
                    break
        return matches


@lru_cache(maxsize=1)
def get_icd_diagnosis_matcher() -> ICDDiagnosisMatcher:
    return ICDDiagnosisMatcher()


def find_similar_diagnoses(diagnosis: str, limit: int = 5) -> List[dict]:
    return get_icd_diagnosis_matcher().find_similar(diagnosis, limit)


def lookup_diagnosis_code(code: str) -> dict | None:
    return get_icd_diagnosis_matcher().lookup_code(code)
