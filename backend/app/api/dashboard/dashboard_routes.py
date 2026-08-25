from __future__ import annotations

import io
import json
import os
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font

from backend.app.core.middleware import require_auth, require_roles
from backend.app.services.workflow_services import (
    audit_service,
    chart_assignment_service,
    risk_service,
)
from backend.app.services.chart_assignment_service import ChartAssignmentError
from backend.app.api.auth.auth_routes import auth_service

bp = Blueprint("dashboard", __name__, url_prefix="/api/dashboard")
chart_service = chart_assignment_service


@bp.route("/coder", methods=["GET"])
@require_auth
@require_roles("coder", "coder_l1", "coder_l2")
def coder_dashboard() -> tuple:
    user = request.current_user
    is_l2 = user.role == "coder_l2"
    return jsonify({
        "success": True,
        "user": {"user_id": user.user_id, "username": user.username, "role": user.role},
        "workflow_stage": "l2" if is_l2 else "l1",
        "claim_bucket": 2 if is_l2 else 1,
        "assigned_chart_ids": [
            chart.chart_id for chart in chart_service.repository.list_charts()
            if chart.assigned_to_user_id == user.user_id
        ],
    }), 200


@bp.route("/admin", methods=["GET"])
@require_auth
@require_roles("admin", "master_admin")
def admin_dashboard() -> tuple:
    return jsonify({
        "success": True,
        "message": "Master admin workspace ready",
        "queue": [chart.chart_id for chart in chart_service.repository.list_charts()],
    }), 200


@bp.route("/supervisor", methods=["GET"])
@require_auth
@require_roles("manager", "supervisor", "admin", "master_admin")
def supervisor_dashboard() -> tuple:
    charts = chart_service.repository.list_charts()
    submissions = risk_service.repository.get_all()
    production_charts_by_user = {}
    quality_rejections_by_user = {}
    for submission in submissions:
        production_charts_by_user.setdefault(submission.user_id, set()).add(submission.chart_id)
        decisions = (submission.user_inputs or {}).get("diagnosis_decisions", {})
        quality_rejections_by_user[submission.user_id] = quality_rejections_by_user.get(submission.user_id, 0) + sum(
            1 for decision in decisions.values() if decision.get("decision") == "rejected"
        )
    return jsonify({
        "success": True,
        "message": "Supervisor workspace ready",
        "buckets": {
            "bucket_1": sum(chart.status in {"queued", "released", "locked", "in_progress"} for chart in charts),
            "bucket_2": sum(chart.status in {"pending_audit", "audit_locked"} for chart in charts),
            "bucket_3": sum(chart.status == "audited" for chart in charts),
        },
        "production_by_user": {
            user_id: len(chart_ids) for user_id, chart_ids in production_charts_by_user.items()
        },
        "quality_rejections_by_user": quality_rejections_by_user,
        "queue": [
            {
                "chart_id": chart.chart_id,
                "original_filename": chart.original_filename,
                "status": chart.status,
                "assigned_to_user_id": chart.assigned_to_user_id,
                "priority": chart.priority,
            }
            for chart in charts
        ],
    }), 200


@bp.route("/manager", methods=["GET"])
@require_auth
@require_roles("manager", "admin", "master_admin")
def manager_dashboard() -> tuple:
    return supervisor_dashboard()


@bp.route("/audit-bucket", methods=["GET"])
@require_auth
@require_roles("manager", "supervisor", "admin", "master_admin")
def audit_bucket() -> tuple:
    """Return the restricted bucket-3 audit projection, without source files or OCR text."""
    users = auth_service.repository
    records = []
    for chart in chart_service.repository.list_charts():
        if chart.status != "audited":
            continue
        latest = risk_service.get_latest_for_chart(chart.chart_id)
        coder = users.get_user_by_id(chart.l1_user_id) if chart.l1_user_id else None
        auditor = users.get_user_by_id(chart.l2_user_id) if chart.l2_user_id else None
        records.append({
            "chart_id": chart.chart_id,
            "patient_details": chart.patient_details,
            "encounter_details": chart.encounter_details,
            "codes": latest.captured_icd10_codes if latest else [],
            "category": chart.category,
            "l1_username": coder.username if coder else None,
            "l2_username": auditor.username if auditor else None,
            "audited_at": chart.updated_at.isoformat() if chart.updated_at else None,
        })
    return jsonify({"success": True, "bucket": 3, "records": records, "count": len(records)}), 200


@bp.route("/export.xlsx", methods=["GET"])
@require_auth
@require_roles("manager", "supervisor", "admin", "master_admin")
def export_coding_workbook():
    workbook = Workbook()
    charts_sheet = workbook.active
    charts_sheet.title = "Charts"
    charts_sheet.append([
        "Chart ID", "Filename", "Status", "Coder", "Coder ID",
        "Submission Status", "Submitted At", "ICD-10 Codes", "RAF Score", "Diagnosis Decisions",
    ])
    codes_sheet = workbook.create_sheet("Codes")
    codes_sheet.append(["Chart ID", "Coder", "ICD-10 Code", "HCC Mappings"])
    decisions_sheet = workbook.create_sheet("Diagnosis Decisions")
    decisions_sheet.append(["Chart ID", "Coder", "Diagnosis", "Decision", "ICD Suggestions"])
    nlp_sheet = workbook.create_sheet("NLP Summary")
    nlp_sheet.append(["Chart ID", "Filename", "Document Type", "NLP Confidence", "Conditions", "OCR Text"])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

    for chart in chart_service.repository.list_charts():
        latest = risk_service.get_latest_for_chart(chart.chart_id)
        is_coder_submission = bool(latest and latest.user_id)
        coder = auth_service.repository.get_user_by_id(latest.user_id) if is_coder_submission else None
        coder_name = coder.username if coder else ""
        inputs = latest.user_inputs if latest else {}
        decisions = inputs.get("diagnosis_decisions", {})
        hcc_mappings = latest.mapped_hcc_versions if latest else []
        charts_sheet.append([
            chart.chart_id,
            chart.original_filename,
            chart.status,
            coder_name,
            latest.user_id if is_coder_submission else "",
            "Submitted" if is_coder_submission else "Not submitted",
            latest.created_at.isoformat() if latest and latest.created_at else "",
            ", ".join(latest.captured_icd10_codes or []) if latest else "",
            latest.calculated_raf_score if latest else "",
            json.dumps(decisions, ensure_ascii=True),
        ])
        for code in latest.captured_icd10_codes if latest else []:
            code_hccs = [item.get("hcc") for item in hcc_mappings if item.get("icd10") == code]
            codes_sheet.append([chart.chart_id, coder_name, code, json.dumps(code_hccs, ensure_ascii=True)])
        for diagnosis, decision in decisions.items():
            decisions_sheet.append([
                chart.chart_id,
                coder_name,
                diagnosis,
                decision.get("decision", ""),
                json.dumps(decision.get("icd10_suggestions", []), ensure_ascii=True),
            ])
        nlp_result = inputs.get("nlp_result", {})
        medical_entities = nlp_result.get("medical_entities", {})
        nlp_sheet.append([
            chart.chart_id,
            chart.original_filename,
            nlp_result.get("document_type", ""),
            nlp_result.get("confidence"),
            ", ".join(medical_entities.get("medical_conditions", [])),
            nlp_result.get("text", ""),
        ])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    os.makedirs("reports", exist_ok=True)
    report_path = os.path.join(
        "reports", f"coding_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    with open(report_path, "wb") as report_file:
        report_file.write(output.getvalue())
    output.seek(0)
    audit_service.record_event(
        action_type="coding_exported",
        entity_type="coding_workbook",
        details={"format": "xlsx", "report_path": report_path},
        user_id=request.current_user.user_id,
    )
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="coding_results.xlsx",
    )


@bp.route("/submit", methods=["POST"])
@require_auth
@require_roles("coder", "coder_l1", "coder_l2")
def submit_coding() -> tuple:
    payload = request.get_json(silent=True) or {}
    chart_id = int(payload.get("chart_id", 0))
    chart = chart_service.repository.get_chart(chart_id)
    if not chart:
        return jsonify({"success": False, "error": "chart_not_found"}), 404
    if chart.assigned_to_user_id != request.current_user.user_id:
        return jsonify({"success": False, "error": "chart_not_assigned_to_coder"}), 403
    try:
        chart.patient_details = payload.get("patient_details", chart.patient_details)
        chart.encounter_details = payload.get("encounter_details", chart.encounter_details)
        chart.category = str(payload.get("category", chart.category)).strip()
        chart = chart_service.submit_chart(
            chart_id=chart_id,
            user_id=request.current_user.user_id,
            actor_role=request.current_user.role,
        )
    except ChartAssignmentError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    record = risk_service.save_submission(
        chart_id=chart_id,
        user_id=request.current_user.user_id,
        user_inputs=payload.get("user_inputs", {}),
        captured_icd10_codes=payload.get("captured_icd10_codes", []),
        mapped_hcc_versions=payload.get("mapped_hcc_versions", []),
        calculated_raf_score=payload.get("calculated_raf_score"),
    )
    audit_service.record_event(
        action_type="coding_submitted",
        entity_type="risk_adjustment_input",
        details={
            "chart_id": chart_id,
            "input_id": record.input_id,
            "status": chart.status,
            "bucket": 2 if chart.status == "pending_audit" else 3,
        },
        user_id=request.current_user.user_id,
        chart_id=chart_id,
        entity_id=str(record.input_id),
    )
    return jsonify({"success": True, "input_id": record.input_id, "chart_status": chart.status}), 200
